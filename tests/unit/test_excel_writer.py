import os
import sys
import tempfile

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "pipeline"))

from excel_writer import COLUMNS, append_features  # noqa: E402

ROW_1 = {
    "transaction_id": "txn-1",
    "terminal_id": "TRM00001",
    "card_bin": "539983",
    "amount_ngn": 5000.0,
    "velocity_1h": 1,
    "geo_jump_km": 0.0,
    "bin_spend_rate": 5000.0,
    "terminal_reversal_count": 0,
    "amount_vs_bin_avg_ratio": 1.0,
    "timestamp": "2026-08-13T09:00:00+00:00",
}
ROW_2 = {**ROW_1, "transaction_id": "txn-2"}


def _new_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    wb.active.append(COLUMNS)
    wb.save(path)


def test_append_features_writes_new_rows():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "features.xlsx")
        _new_workbook(path)

        append_features([ROW_1, ROW_2], xlsx_path=path)

        ws = openpyxl.load_workbook(path).active
        ids = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)]
        assert ids == ["txn-1", "txn-2"]


def test_append_features_skips_replayed_transaction_ids():
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "features.xlsx")
        _new_workbook(path)

        append_features([ROW_1], xlsx_path=path)
        # Simulate a Spark foreachBatch replay: txn-1 arrives again alongside
        # one genuinely new row.
        append_features([ROW_1, ROW_2], xlsx_path=path)

        ws = openpyxl.load_workbook(path).active
        ids = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)]
        assert ids == ["txn-1", "txn-2"]
