"""Offline pandas equivalent of feature_windows.py's Spark transforms.

spark_consumer.py computes velocity_1h/geo_jump_km/bin_spend_rate/
terminal_reversal_count as a live Spark Structured Streaming job reading
from Kafka - the intended production path, but unusable here (no Docker/
Spark/Kafka available in this environment, see SCRUM-20). This script
recomputes the exact same four window definitions in pandas, directly
from transactions_raw.xlsx, so data/features.xlsx can be backfilled for
bulk-generated historical volume (see producer/bulk_generate.py) without
a running streaming cluster.

Window semantics are matched deliberately to feature_windows.py:
    - velocity_1h, bin_spend_rate: trailing 1h window, INCLUSIVE of both
      endpoints (Spark's rangeBetween(-3600, 0) in seconds).
    - terminal_reversal_count: trailing 24h window, same inclusivity.
    - geo_jump_km: haversine distance to the immediately preceding
      transaction for the same terminal_id, 0.0 if there is none - reuses
      geo.haversine_km, which is already dependency-free (no pyspark).

Implementation note: trailing windows are computed with a searchsorted
two-pointer per group (_trailing_window_agg) rather than
pandas' groupby().rolling(on=...), because that returns a
(group_key, timestamp) MultiIndex that raises on duplicate timestamps -
and with tens of thousands of rows generated back-to-back, duplicate
millisecond-resolution timestamps within the same terminal/card_bin are
common. One simplification versus Spark's exact RANGE-frame semantics:
ties (two rows for the same group at the identical timestamp) are broken
by row position rather than treated as mutual peers, i.e. an earlier row
does not "see" a later row with the exact same timestamp. This only
affects exact-timestamp collisions and is judged acceptable for a batch
backfill utility.

Recomputes features for the FULL transactions_raw.xlsx (not just newly
appended rows) and overwrites features.xlsx, since the rolling windows
are only correct when computed over each terminal's/card_bin's complete
chronological history.
"""
import numpy as np
import openpyxl
import pandas as pd
from filelock import FileLock

from geo import haversine_km

TRANSACTIONS_XLSX = "data/transactions_raw.xlsx"
FEATURES_XLSX = "data/features.xlsx"

FEATURE_COLUMNS = [
    "transaction_id",
    "terminal_id",
    "card_bin",
    "amount_ngn",
    "velocity_1h",
    "geo_jump_km",
    "bin_spend_rate",
    "terminal_reversal_count",
    "amount_vs_bin_avg_ratio",
    "timestamp",
]


def load_transactions(xlsx_path: str = TRANSACTIONS_XLSX) -> pd.DataFrame:
    lock_path = f"{xlsx_path}.lock"
    with FileLock(lock_path, timeout=60):
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    # format="ISO8601" tolerates the mix of with/without-microseconds timestamps
    # produced by different transaction_generator.py runs.
    df["timestamp"] = pd.to_datetime(df["timestamp"], utc=True, format="ISO8601")
    return df


def _trailing_window_agg(
    df: pd.DataFrame, group_col: str, value_col: str | None, window_seconds: int, agg: str
) -> np.ndarray:
    """Trailing, inclusive-both-ends window aggregate of value_col (or a
    constant 1 if value_col is None, for counting) grouped by group_col.
    df must already be sorted by timestamp ascending. agg is "sum" or "mean"."""
    n = len(df)
    result = np.zeros(n, dtype=float)
    ts = df["timestamp"].to_numpy(dtype="datetime64[ns]")
    vals = np.ones(n, dtype=float) if value_col is None else df[value_col].to_numpy(dtype=float)
    window_td = np.timedelta64(window_seconds, "s")

    for idx in df.groupby(group_col).indices.values():
        idx = np.asarray(idx)
        g_ts = ts[idx]
        g_vals = vals[idx]
        left_bound = g_ts - window_td
        start = np.searchsorted(g_ts, left_bound, side="left")
        prefix = np.concatenate(([0.0], np.cumsum(g_vals)))
        positions = np.arange(len(idx))
        window_sum = prefix[positions + 1] - prefix[start]
        if agg == "mean":
            window_count = positions - start + 1
            result[idx] = window_sum / window_count
        else:
            result[idx] = window_sum

    return result


def _velocity_1h(df: pd.DataFrame) -> np.ndarray:
    """Trailing-1h transaction count per terminal_id - mirrors
    feature_windows.with_velocity_1h."""
    return _trailing_window_agg(df, "terminal_id", None, 3600, "sum")


def _geo_jump_km(df: pd.DataFrame) -> pd.Series:
    """Haversine distance to the previous transaction on the same
    terminal_id, 0.0 for a terminal's first transaction - mirrors
    feature_windows.with_geo_jump."""
    prev_lat = df.groupby("terminal_id")["latitude"].shift(1)
    prev_lon = df.groupby("terminal_id")["longitude"].shift(1)
    has_prev = prev_lat.notna()

    jump = pd.Series(0.0, index=df.index)
    jump.loc[has_prev] = [
        haversine_km(pl, po, la, lo)
        for pl, po, la, lo in zip(
            prev_lat[has_prev], prev_lon[has_prev], df.loc[has_prev, "latitude"], df.loc[has_prev, "longitude"]
        )
    ]
    return jump


def _bin_spend_rate(df: pd.DataFrame) -> np.ndarray:
    """Trailing-1h rolling average amount_ngn per card_bin - mirrors
    feature_windows.with_bin_spend_rate."""
    return _trailing_window_agg(df, "card_bin", "amount_ngn", 3600, "mean")


def _terminal_reversal_count(df: pd.DataFrame) -> np.ndarray:
    """Trailing-24h count of fake_reversal-flagged transactions per
    terminal_id - mirrors feature_windows.with_terminal_reversal_count."""
    df = df.assign(_is_reversal=(df["fraud_type"] == "fake_reversal").astype(float))
    return _trailing_window_agg(df, "terminal_id", "_is_reversal", 86400, "sum")


def build_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.sort_values("timestamp").reset_index(drop=True)

    bin_spend_rate = _bin_spend_rate(df)
    features = pd.DataFrame(
        {
            "transaction_id": df["transaction_id"],
            "terminal_id": df["terminal_id"],
            "card_bin": df["card_bin"],
            "amount_ngn": df["amount_ngn"],
            "velocity_1h": _velocity_1h(df),
            "geo_jump_km": _geo_jump_km(df),
            "bin_spend_rate": bin_spend_rate,
            "terminal_reversal_count": _terminal_reversal_count(df),
            # Sharper anomaly signal than amount_ngn or bin_spend_rate alone -
            # see feature_windows.with_amount_vs_bin_avg_ratio, which this mirrors.
            "amount_vs_bin_avg_ratio": df["amount_ngn"].astype(float) / bin_spend_rate,
            "timestamp": df["timestamp"],
        }
    )
    return features[FEATURE_COLUMNS]


def write_features(features: pd.DataFrame, xlsx_path: str = FEATURES_XLSX) -> None:
    lock_path = f"{xlsx_path}.lock"
    with FileLock(lock_path, timeout=60):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(FEATURE_COLUMNS)
        for row in features.itertuples(index=False):
            ws.append(
                [
                    row.transaction_id,
                    row.terminal_id,
                    row.card_bin,
                    row.amount_ngn,
                    row.velocity_1h,
                    row.geo_jump_km,
                    row.bin_spend_rate,
                    row.terminal_reversal_count,
                    row.amount_vs_bin_avg_ratio,
                    row.timestamp.isoformat(),
                ]
            )
        wb.save(xlsx_path)


def main() -> None:
    transactions = load_transactions()
    features = build_features(transactions)
    write_features(features)
    print(f"Recomputed features for {len(features)} transactions -> {FEATURES_XLSX}")


if __name__ == "__main__":
    main()
