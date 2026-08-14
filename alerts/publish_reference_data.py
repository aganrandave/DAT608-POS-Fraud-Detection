"""One-shot bootstrap: publishes data/reference/terminals.xlsx and
merchants.xlsx to the terminal-reference and merchant-reference Kafka
topics, so alerts/tables/terminal_reference.sql and merchant_reference.sql
have real data to back 03_create_alert_stream.sql's joins.

Previously nothing published these topics at all - the ksqlDB tables would
register successfully (CREATE TABLE just declares a schema over a topic)
but stay permanently empty, so every alert's terminal_name/merchant_name
would come back null.

Run once per fresh Kafka cluster, before the reference tables are created -
see alerts/README.md and the alerts container's entrypoint.sh. Not a
long-running service like kafka_consumer.py.
"""
import json
import os
import time

import openpyxl
from kafka import KafkaProducer
from kafka.errors import NoBrokersAvailable

KAFKA_BOOTSTRAP_SERVERS = os.getenv("KAFKA_BOOTSTRAP_SERVERS", "localhost:9092")
TERMINALS_XLSX = os.getenv("TERMINALS_XLSX", "data/reference/terminals.xlsx")
MERCHANTS_XLSX = os.getenv("MERCHANTS_XLSX", "data/reference/merchants.xlsx")

CONNECT_RETRIES = 12
CONNECT_RETRY_DELAY_SECONDS = 5


def read_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        return [dict(zip(header, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    finally:
        # read_only workbooks keep the underlying file open until closed -
        # left unclosed, this leaks a handle every run and (on Windows)
        # blocks the file from being deleted or rewritten afterward.
        wb.close()


def build_producer() -> KafkaProducer:
    """Mirrors producer/kafka_producer.py's retry loop - this script runs
    at container startup, right after Kafka's port opens, same race as
    every other service that connects at boot."""
    last_error: NoBrokersAvailable | None = None
    for attempt in range(1, CONNECT_RETRIES + 1):
        try:
            return KafkaProducer(
                bootstrap_servers=KAFKA_BOOTSTRAP_SERVERS,
                value_serializer=lambda v: json.dumps(v).encode("utf-8"),
                key_serializer=lambda k: k.encode("utf-8"),
            )
        except NoBrokersAvailable as exc:
            last_error = exc
            print(f"Kafka not reachable yet (attempt {attempt}/{CONNECT_RETRIES}), retrying...")
            time.sleep(CONNECT_RETRY_DELAY_SECONDS)
    raise last_error


def publish(producer: KafkaProducer, topic: str, key_field: str, rows: list[dict]) -> int:
    for row in rows:
        producer.send(topic, key=str(row[key_field]), value=row)
    producer.flush()
    return len(rows)


def run() -> None:
    producer = build_producer()
    terminal_count = publish(producer, "terminal-reference", "terminal_id", read_rows(TERMINALS_XLSX))
    merchant_count = publish(producer, "merchant-reference", "merchant_id", read_rows(MERCHANTS_XLSX))
    print(f"Published {terminal_count} terminals, {merchant_count} merchants")


if __name__ == "__main__":
    run()
